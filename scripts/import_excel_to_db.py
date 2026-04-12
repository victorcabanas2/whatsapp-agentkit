"""
Script para importar contactos del Excel a la base de datos del bot.
Genera números de teléfono ficticios y calcula scores automáticos.
"""

import asyncio
import openpyxl
from datetime import datetime, timedelta
import random
import sys
import os

# Agregar agent al path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from agent.memory import inicializar_db, async_session, Lead
from sqlalchemy import select


async def generar_telefono():
    """Genera un teléfono ficticio de Paraguay (+595...)."""
    # Formatos válidos: +595 9XX XXXXXX (celular)
    area = random.choice(['91', '92', '93', '94', '95', '96', '97', '98', '99'])
    numero = ''.join([str(random.randint(0, 9)) for _ in range(7)])
    return f"+595{area}{numero}"


async def cargar_excel():
    """Carga los contactos del Excel a la base de datos."""

    # Inicializar DB
    await inicializar_db()
    print("✓ Base de datos inicializada")

    # Leer Excel
    wb = openpyxl.load_workbook(
        "/Users/victorcabanas/whatsapp-agentkit/knowledge/clientes rebody importado.xlsx"
    )
    ws = wb.active

    contactos = []
    telefonos_existentes = set()

    # Verificar qué telefonos ya existen
    async with async_session() as session:
        resultado = await session.execute(select(Lead.telefono))
        telefonos_existentes = set(resultado.scalars().all())

    print(f"✓ {len(telefonos_existentes)} telefonos ya en DB")

    # Leer contactos únicos
    nombres_vistos = set()
    for row in range(2, ws.max_row + 1):
        nombre = ws.cell(row, 1).value
        if nombre and isinstance(nombre, str):
            nombre_limpio = nombre.strip()
            if nombre_limpio and nombre_limpio not in nombres_vistos:
                nombres_vistos.add(nombre_limpio)
                contactos.append(nombre_limpio)

    print(f"✓ {len(contactos)} contactos únicos en Excel")

    # Registrar contactos
    registrados = 0
    duplicados = 0

    for i, nombre in enumerate(contactos):
        # Generar teléfono único
        while True:
            telefono = await generar_telefono()
            if telefono not in telefonos_existentes:
                telefonos_existentes.add(telefono)
                break

        # Asignar score y estado aleatorios basado en "actividad"
        # (simulamos más activos con mejor score)
        if i % 5 == 0:  # 20% hot
            score = random.randint(70, 95)
            intencion = "hot"
        elif i % 3 == 0:  # 33% warm
            score = random.randint(50, 69)
            intencion = "warm"
        else:  # 47% cold
            score = random.randint(20, 49)
            intencion = "cold"

        # Generar fechas aleatorias (últimos 90 días)
        dias_atras = random.randint(1, 90)
        primer_contacto = datetime.utcnow() - timedelta(days=dias_atras)
        ultimo_mensaje = primer_contacto + timedelta(hours=random.randint(1, 500))

        try:
            async with async_session() as session:
                # Crear lead
                lead = Lead(
                    telefono=telefono,
                    nombre=nombre,
                    primer_contacto=primer_contacto,
                    ultimo_mensaje=ultimo_mensaje,
                    score=score,
                    intencion=intencion,
                    urgencia=random.choice(['baja', 'media', 'alta']),
                    fue_cliente=random.random() < 0.15,  # 15% fueron clientes
                )
                session.add(lead)
                await session.commit()
                registrados += 1
        except Exception as e:
            duplicados += 1
            if i % 50 == 0:
                print(f"  [{i}/{len(contactos)}] Procesando...")

    print(f"\n✅ Importación completada:")
    print(f"   - Registrados: {registrados}")
    print(f"   - Duplicados: {duplicados}")
    print(f"   - Total en DB: {registrados + len(telefonos_existentes) - registrados}")


if __name__ == "__main__":
    asyncio.run(cargar_excel())
