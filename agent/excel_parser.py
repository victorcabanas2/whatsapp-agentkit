# agent/excel_parser.py — Parser para importar clientes de Excel

"""
Funciones para leer archivos Excel con clientes y cargarlos en la base de datos.
"""

import os
import logging
from agent.memory import marcar_cliente_como_antiguo, async_session, Lead
from sqlalchemy import select

logger = logging.getLogger("agentkit")


async def cargar_clientes_desde_excel(ruta_archivo: str) -> dict:
    """
    Lee un archivo Excel con clientes antiguos y los carga en la BD.

    Espera que el Excel tenga las columnas:
    - Teléfono (o Phone, celular, etc.)
    - Nombre (o Name, Cliente, etc.)
    - Productos Comprados (o Productos, Purchase History, etc.)
    - Última Compra (o Last Purchase, Última Compra, etc.)

    Returns:
        {
            "exitosos": int,
            "errores": int,
            "duplicados": int,
            "total": int,
            "detalles": []
        }
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl no instalado. Instala con: pip install openpyxl")
        return {"error": "openpyxl no instalado"}

    if not os.path.exists(ruta_archivo):
        return {"error": f"Archivo no encontrado: {ruta_archivo}"}

    exitosos = 0
    errores = 0
    duplicados = 0
    detalles = []

    try:
        # Cargar el libro
        wb = openpyxl.load_workbook(ruta_archivo, data_only=True)
        ws = wb.active

        # Detectar columnas por nombre (primero si existen, sino por índice)
        headers = [cell.value for cell in ws[1]]
        logger.info(f"Headers detectados: {headers}")

        # Mapeo de columnas — buscar variaciones
        def encontrar_columna(nombres_posibles):
            """Busca una columna por múltiples nombres posibles (case-insensitive)."""
            for i, header in enumerate(headers):
                if header and header.lower().strip() in [n.lower() for n in nombres_posibles]:
                    return i
            return None

        col_telefono = encontrar_columna(["teléfono", "telefono", "phone", "celular", "número", "whatsapp"])
        col_nombre = encontrar_columna(["nombre", "name", "cliente", "cliente name", "nombre cliente"])
        col_productos = encontrar_columna(["productos comprados", "productos", "purchase history", "compras", "historial"])
        col_ultima_compra = encontrar_columna(["última compra", "ultima compra", "last purchase", "fecha última compra"])

        if col_telefono is None:
            return {"error": "No se encontró columna de teléfono"}

        logger.info(f"Columnas detectadas: tel={col_telefono}, name={col_nombre}, prod={col_productos}, fecha={col_ultima_compra}")

        # Procesar filas (saltar header)
        async with async_session() as session:
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Extraer valores
                    telefono = str(row[col_telefono]).strip() if col_telefono is not None and row[col_telefono] else None
                    nombre = str(row[col_nombre]).strip() if col_nombre is not None and row[col_nombre] else "Sin nombre"
                    productos = str(row[col_productos]).strip() if col_productos is not None and row[col_productos] else ""
                    ultima_compra = str(row[col_ultima_compra]).strip() if col_ultima_compra is not None and row[col_ultima_compra] else ""

                    # Validar teléfono
                    if not telefono or telefono == "None":
                        detalles.append({"fila": row_idx, "error": "Teléfono vacío"})
                        errores += 1
                        continue

                    # Limpiar teléfono (remover caracteres especiales)
                    telefono_limpio = telefono.replace("+", "").replace(" ", "").replace("-", "")

                    # Verificar si ya existe
                    query = select(Lead).where(Lead.telefono == telefono_limpio)
                    result = await session.execute(query)
                    lead_existente = result.scalar_one_or_none()

                    if lead_existente and lead_existente.es_cliente_previo:
                        duplicados += 1
                        detalles.append({"fila": row_idx, "telefono": telefono_limpio, "status": "ya existe"})
                        continue

                    # Guardar como cliente antiguo
                    success = await marcar_cliente_como_antiguo(
                        telefono=telefono_limpio,
                        nombre=nombre,
                        productos_comprados=productos,
                        historial_resumen=f"Última compra: {ultima_compra}" if ultima_compra else ""
                    )

                    if success:
                        exitosos += 1
                        detalles.append({"fila": row_idx, "telefono": telefono_limpio, "nombre": nombre, "status": "✓ Importado"})
                        logger.info(f"✓ Cliente importado: {nombre} ({telefono_limpio})")
                    else:
                        errores += 1
                        detalles.append({"fila": row_idx, "telefono": telefono_limpio, "error": "Error al guardar"})

                except Exception as e:
                    errores += 1
                    detalles.append({"fila": row_idx, "error": str(e)})
                    logger.error(f"Error procesando fila {row_idx}: {e}")

        wb.close()

        return {
            "exitosos": exitosos,
            "errores": errores,
            "duplicados": duplicados,
            "total": exitosos + errores,
            "detalles": detalles
        }

    except Exception as e:
        logger.error(f"Error al cargar Excel: {e}")
        return {"error": str(e)}


async def importar_clientes_de_archivo_default():
    """
    Intenta cargar el archivo default de clientes importados.
    Archivo: knowledge/clientes rebody importado.xlsx
    """
    archivo = "knowledge/clientes rebody importado.xlsx"
    if os.path.exists(archivo):
        logger.info(f"Cargando clientes desde {archivo}...")
        return await cargar_clientes_desde_excel(archivo)
    else:
        logger.warning(f"Archivo de clientes no encontrado: {archivo}")
        return {"error": "Archivo no encontrado"}
